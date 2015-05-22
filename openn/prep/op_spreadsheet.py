#!/usr/bin/env python
# -*- coding: utf-8 -*-

import sys
import simplejson
import re

from copy import deepcopy

from openn.prep import langs
from openn.openn_exception import OPennException

from openpyxl import load_workbook
from openpyxl.workbook import Workbook


class OPSpreadsheet:

    MIN_YEAR = -5000
    MAX_YEAR = 3000

    # conversion of John Gruber's URI RE adapted from:
    #   https://gist.github.com/uogbuji/705383
    URI_RE = re.compile(ur'(?i)\b((?:[a-z][\w-]+:(?:/{1,3}|[a-z0-9%])|www\d{0,3}[.]|[a-z0-9.\-]+[.][a-z]{2,4}/)(?:[^\s()<>]+|\(([^\s()<>]+|(\([^\s()<>]+\)))*\))+(?:\(([^\s()<>]+|(\([^\s()<>]+\)))*\)|[^\s`!()\[\]{};:\'".,<>?\xab\xbb\u201c\u201d\u2018\u2019]))')

    YEAR_RE = re.compile(ur'^-?\d{1,4}$')
    # this an imperfect regex for this purpose and can only be assumed
    # to match most email address
    # It's taken from here:
    #
    #   http://www.regular-expressions.info/email.html
    #
    # For the ugly details on regexes and email address, see this discussion:
    #
    #   http://stackoverflow.com/questions/201323/using-a-regular-expression-to-validate-an-email-address
    #
    EMAIL_RE = re.compile(r'\b[A-Z0-9._%+-]+@[A-Z0-9.-]+\.[A-Z]{2,4}\b', re.I)

    # The offset to use to find the first value column; the data
    # starts 2 columns past the column the headering is in:
    #
    # Field          | Req't | Entry 1
    # ---------------|-------|--------
    # Call number    |   R   | MS ABC 123
    #
    FIELD_COLUMN_OFFSET = 2

    ######################################################################
    # Static methods
    ######################################################################

    @staticmethod
    def is_valid_uri(val):
        return OPSpreadsheet.URI_RE.match(val) is not None

    @staticmethod
    def is_valid_year(val):
        return OPSpreadsheet.YEAR_RE.match(str(val)) and \
            int(val) in xrange(OPSpreadsheet.MIN_YEAR, OPSpreadsheet.MAX_YEAR + 1)

    @staticmethod
    def is_valid_email(val):
        return OPSpreadsheet.EMAIL_RE.match(val) is not None

    @staticmethod
    def is_valid_lang(val):
        return langs.is_valid_lang(val)


    ######################################################################
    # Instance methods
    ######################################################################

    def __init__(self, xlsx_file, config):
        """Create a new OPSpreadsheet and find all description sheet headings.

        """
        self.config    = deepcopy(config)
        self.xlsx_path = xlsx_file
        self.workbook  = load_workbook(self.xlsx_path)
        self.errors    = []
        self.warnings  = []
        self._set_headings()

    # --------------------------------------------------------------------
    # Properties
    # --------------------------------------------------------------------

    @property
    def required_fields(self):
        """Return details dict for all fields where 'required' is True.

        """
        return [ x for x in self.fields.itervalues() if x['required'] ]

    @property
    def description_sheet(self):
	return self.workbook.get_sheet_by_name('Description')

    @property
    def fields(self):
        return self.config['fields']

    # --------------------------------------------------------------------
    # Validation
    # --------------------------------------------------------------------

    def has_description_errors(self):
        return len(self.errors) > 0

    def has_description_warnings(self):
        return len(self.warnings) > 0

    def validate_description(self):
        self.check_required_headings()
        self.check_description_values()

    def check_required_headings(self):
        for field in self.required_fields:
            if not field['locus']:
                msg = 'Required field is missing: %s' % (field['field_name'],)
                self.errors.append(msg)

    def check_description_values(self):
        for attr in self.fields:
            self.validate_field(attr)

    def validate_blank(self, attr):
        """If 'attr' has a blank rule and field is not blank, validate
        whether field may have a value.

        """
        if (self.blank_rule(attr) is None or self.is_blank(attr)):
            return

        blankrule  = self.fields[attr]['blank']
        ifclause   = blankrule['if']
        other_attr = ifclause['field']
        condition  = ifclause['is']

        if condition == 'BLANK':
            if self.is_blank(other_attr):
                msg = '"%s" must be blank if "%s" is blank; found: %s' % (
                    self.field_name(attr), self.field_name(other_attr),
                    ', '.join(self._values_quoted(attr)))
                self.errors.append(msg)
        elif condition == 'NONBLANK':
            if self.is_present(other_attr):
                msg = '"%s" must be blank if "%s" has a value; found: %s' % (
                    self.field_name(attr), self.field_name(other_attr),
                    ', '.join(self._values_quoted(attr)))
                self.errors.append(msg)
        elif isinstance(condition,list):
            for val in self.values(other_attr):
                if val in condition and self.is_present(attr):
                    msg = '"%s" must be blank if "%s" is "%s"; found: %s' % (
                        self.field_name(attr), self.field_name(other_attr), val,
                        ', '.join(self._values_quoted(attr)))
                    self.errors.append(msg)
        else:
            raise OPennException('Unknown condition type: "%s"' % (condition,))

    def validate_conditional(self, attr, required):
        """Validate a conditional requirement rule."""
        ifclause    = required['if']
        other_attr  = ifclause['field']
        condition   = ifclause['is']
        if condition == 'BLANK':
            if self.is_blank(other_attr):
                if self.is_blank(attr):
                    msg = '"%s" cannot be blank if "%s" is blank' % (
                        self.field_name(attr), self.field_name(other_attr))
                    self.errors.append(msg)
        elif condition == 'NONBLANK':
            if self.is_present(other_attr):
                if self.is_blank(attr):
                    msg = '"%s" cannot be blank if "%s" has a value' % (
                        self.field_name(attr), self.field_name(other_attr))
                    self.errors.append(msg)
        elif isinstance(condition, list):
            for val in self.values(other_attr):
                if val in condition and self.is_blank(attr):
                    msg = '"%s" cannot be blank if "%s" is "%s"' % (
                        self.field_name(attr), self.field_name(other_attr), val)
                    self.errors.append(msg)
        else:
            raise OPennException("Unknown condition type: %s" % (condition,))

    def validate_requirement(self, attr):
        """Validate required or conditionally required field if 'required' is
        present and non-False.

        """
        details = self.fields[attr]
        field_name = details['field_name']
        required = details['required']
        # print required
        if required == True and self.is_blank(attr):
            msg = '%s cannot be blank' % (details['field_name'],)
            self.errors.append(msg)
        elif isinstance(required, dict):
            self.validate_conditional(attr, required)

    def validate_value_list(self, attr):
        """Validate field whose values must come from a list if 'value_list'
        present.

        """
        value_list = self.value_list(attr)
        if value_list is None: return

        for val in self.values(attr):
            if val not in value_list:
                msg = '"%s" value "%s" not valid; expected one of: %s' % (
                    self.field_name(attr), val, ', '.join(self._list_quoted(value_list)))
                self.errors.append(msg)

    def validate_repeating(self, attr):
        """Validate non-repeating fields if 'repeating' set to False.

        """
        if not self.repeating(attr) and len(self.values(attr)) > 1:
            msg = "More than one value found in non-repeating field %s: %s" % (
                self.field_name(attr), ', '.join(self._values_quoted(attr)))
            self.errors.append(msg)

    def validate_field(self, attr):
        """Perform all validations for field."""
        # first see if the field is missing
        if self.is_field_missing(attr): return

        self.validate_requirement(attr)
        self.validate_blank(attr)
        self.validate_value_list(attr)
        self.validate_repeating(attr)
        self.validate_data_type(attr)

    def validate_data_type(self, attr):
        """Validate all values for field against configured 'data_type'.

        """
        data_type = self.data_type(attr)
        field_name = self.field_name(attr)
        for val in self.values(attr):
            self._do_type_validation(field_name, val, self.data_type(attr))

    # --------------------------------------------------------------------
    # Field accessors
    # --------------------------------------------------------------------

    def is_blank(self, field):
        """Return True if the field has no value present."""
        values = self.values(field)
        return len(values) == 0

    def is_present(self, field):
        """Return True if the field has one or more values."""
        return not self.is_blank(field)

    def is_field_missing(self, attr):
        """Return True if the field is not on the description sheet."""
        return self.locus(attr) is None

    def values(self, attr):
        """Return all values for attr's field."""
        details = self.fields[attr]
        if details.get('cell_values') is None:
            details['cell_values'] = self._extract_values(attr)
        return details.get('cell_values')

    def field_name(self, attr):
        """Return the 'field_name' for attr's field."""
        if self.fields.get(attr):
            return self.fields[attr]['field_name']

    def locus(self, attr):
        """Return the header 'locus' for attr's field."""
        if self.fields.get(attr):
            return self.fields[attr].get('locus')

    def value_list(self, attr):
        """If present, return the 'value_list' for attr's field."""
        if self.fields.get(attr):
            return self.fields[attr].get('value_list')

    def requirement(self, attr):
        """Return the 'require' value for attr's field."""
        if self.fields.get(attr):
            return self.fields[attr].get('required')

    def blank_rule(self, attr):
        """If present, return the 'blank' rule for attr's field."""
        if self.fields.get(attr):
            return self.fields[attr].get('blank')

    def repeating(self, attr):
        """Return the 'repeating' value for attr's field."""
        if self.fields.get(attr):
            return self.fields[attr].get('repeating')

    def data_type(self, attr):
        """Return the 'data_type' for attr's field."""
        if self.fields.get(attr):
            return self.fields[attr].get('data_type')

    ######################################################################
    # _ methods
    ######################################################################

    def _values_quoted(self, attr, q='"'):
        return self._list_quoted(self.values(attr), q=q)

    def _list_quoted(self, strings, q='"'):
        return [ "%s%s%s" % (q,s,q) for s in strings ]

    def _normalize(self, s):
        if s is not None:
            return re.sub(r'\W+', '', str(s)).lower()

    def _do_type_validation(self, field, value, data_type):
        if data_type == 'year':
            if not OPSpreadsheet.is_valid_year(value):
                self.errors.append(self._format_error(field, value, data_type))
        elif data_type == 'uri':
            if not OPSpreadsheet.is_valid_uri(value):
                self.errors.append(self._format_error(field, value, data_type))
        elif data_type == 'lang':
            if not OPSpreadsheet.is_valid_lang(value):
                self.errors.append(self._format_error(field, value, data_type))
        elif data_type == 'email':
            if not OPSpreadsheet.is_valid_email(value):
                self.warnings.append(
                    self._format_error(field, value, data_type))
        elif data_type == 'string':
            pass
        else:
            raise OPennException('Unknown data type: "%s"' % (data_type,))

    def _set_headings(self):
        """Find and set the headings locus for each field in config."""
        for attr in self.fields:
            details          = self.fields[attr]
            field_name       = details['field_name']
            locus            = self._find_heading_locus(field_name)
            details['locus'] = locus

    def _find_heading_locus(self, field_name):
        """Find the heading locus for 'field_name'. Field name is compared to
        cell values by normalizing both strings and comparing.
        Normalization removes all non-word characaters and converts
        the remaining characters to lower case.

        """
        locus = []
        sheet = self.description_sheet
        for row in xrange(1, sheet.max_row+1):
            for col in xrange(1, 21):
                cell = self.description_sheet.cell(column=col, row=row)
                value = cell.value if cell is not None else ''
                if self._normalize(value) == self._normalize(field_name):
                    return {'col': col,'row': row }

    def _extract_values(self, attr):
        """Based on the locus of attr's field's header."""
        if self.locus(attr) is None: return

        locus   = self.locus(attr)
        row     = locus['row']
        # read the first 20 columns past the heading locus
        data_col = locus['col'] + self.FIELD_COLUMN_OFFSET

        vals = []
        for col in xrange(data_col, data_col + 21):
            cell = self.description_sheet.cell(column=col,row=row)
            if cell.value is not None and str(cell.value).strip() != '':
                vals.append(cell.value)
            else:
                vals.append(None)

        # strip off trailing Nones
        while len(vals) > 0 and vals[-1] is None:
            del vals[-1]

        return vals

    def _format_error(self, field_name, value, data_type):
        return "%s is not a valid %s: %s" % (field_name, data_type, value)
