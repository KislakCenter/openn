import sys

from openpyxl import load_workbook
from openpyxl.workbook import workbook

wkbk = load_workbook(sys.argv[1])

wksh = wkbk.get_sheet_by_name(sys.argv[2])

print '===  max column ==='
print wksh.get_highest_column()
print wksh.max_column

print '\n===  max row ==='
print wksh.get_highest_row()
print wksh.max_row
